""".xlsx parser (openpyxl, MIT) -- news cuesheets and (later) schedules.

Returns every sheet as a table of stringified cells; column semantics are
resolved in normalize/cuesheet.py so odd layouts only touch that one place.
"""
from .base import BaseParser, RawDoc


class XlsxParser(BaseParser):
    def parse(self, path: str, hint: str | None = None) -> RawDoc:
        from openpyxl import load_workbook  # lazy: xlsx-reading workers only

        wb = load_workbook(path, read_only=True, data_only=True)
        tables, titles = [], []
        for ws in wb.worksheets:
            rows = []
            for row in ws.iter_rows(values_only=True):
                cells = ["" if v is None else str(v).strip() for v in row]
                if any(cells):
                    rows.append(cells)
            if rows:
                tables.append(rows)
                titles.append(ws.title)
        wb.close()
        return RawDoc(source_path=path, source_format="xlsx",
                      doc_type_hint="cuesheet",
                      tables=tables, meta={"sheets": titles})
